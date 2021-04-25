# from openpyxl import Workbook

# wb = Workbook()
# wb.save('我的工作薄.xlsx')

import openpyxl

wb = openpyxl.Workbook()
for m in range(1, 13):
    wb.create_sheet('%d 月' % m)
wb.remove(wb['Sheet'])
wb.save('2021年计划表.xlsx')